import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
#This script encodes user's input into color resistance representation.
#Please ensure to create the file 'Flag.xlsx' before running this script!!!
#I tried creating it with python, but always get errors and the file is corrupted!


#This function processes user's input and convert it into decimal representation
def input_processing(flag):
    split_flag = list(flag)
    flag = ""

    for index in range(len(split_flag)):
        if index < len(split_flag) - 1:
            flag += str(ord(split_flag[index])) + "//"
        else:
            flag += str(ord(split_flag[index])) + "END"

    return list(flag)

#This function convert decimals to colors and return them in a list
def get_colors(processed_input, COLORS_TABLE):
    colors = []
    for char in processed_input:
        if (char != "/") and (char != "E"):
            colors.append(COLORS_TABLE.get(char))
        elif char == "/":
            colors.append(char)
        else:
            colors.append("END")
            break
    return colors

#This function fills cells with colors
def fill_cells(sheet, colors, column):
    for color in colors:
        active_cell = sheet.cell(row=1, column=column)
        if (color != "/")  and (color != "END"):
        # Create a fill pattern
            fill = PatternFill(start_color=color, fill_type="solid")
            active_cell.fill = fill
            column+=1
        elif (color == "/") and (active_cell.value == None):
            active_cell.value = color
        elif (color == "/") and (active_cell.value != None):
            active_cell.value += color
            active_cell.alignment = Alignment(horizontal= "center")
            active_cell.font = Font(bold=True)
            column+=1
        else:
            active_cell.value = "END"
            active_cell.alignment = Alignment(horizontal= "center")
            active_cell.font = Font(bold=True)
            break


#This function setups COLORS_TABLE and workbook
def initialization():
    return {"0":"000000", "1":"C00000", "2":"FF0000", "3":"FCA904", "4":"FFFF00", "5":"00B050", "6":"0070C0", "7":"CC99FF", "8":"808080", "9":"FFFFFF"} 

COLORS_TABLE= initialization()

workbook = openpyxl.load_workbook("Flag.xlsx")
active_worksheet = workbook.active

#test input = ingeneer{I_C0u1dnT_R3s1s7}

flag = input("Type your flag to encode to color resistance:")
processed_input = input_processing(flag)
cell_colors = get_colors(processed_input, COLORS_TABLE)
column_number = 1
fill_cells(active_worksheet, cell_colors, column_number)


# Save the workbook
workbook.save("Flag.xlsx")
print("Your message has been encoded successfully ;)\nPlease check the file: Flag.xlsx")
