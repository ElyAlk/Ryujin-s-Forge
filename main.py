from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

#from openpyxl.drawing.image import Image
# Data Initialization
print("Initalizaing data...")
today = datetime.now().today()
week_days = {0: 'الأثنين', 1: 'الثلاثاء', 2: 'الأربعاء', 3: 'الخميس', 4: 'الجمعة', 5: 'السبت', 6: 'الأحد'}
dated_days = []
template_file_path = 'assets/Attendance Sheet - Template.xlsx'
output_file_path = 'C:/Users/Elyas/Desktop/Eng Elyas/الموظفين/سجلات الحضور/سجل حضور الموظفين - ' + today.strftime('%d-%m-%Y') + '.xlsx'
image_path = 'assets/ETCSA-LOGO.png'
title_cell = 'A8'
colored_side = Side(style='thin', color='FFFFFF')
current_rows = 10
current_columns = 3

for day in week_days:
    day
    future_date = today + timedelta(days=day)
    dated_days.append(week_days[future_date.weekday()] + " " + future_date.strftime("%d/%m"))
print(dated_days)
start_work_day = dated_days[0]
end_work_day= dated_days[5]
print('Data Initialized successfully.')

# Opening the attendance template
workbook = load_workbook(template_file_path, data_only=False)
worksheet = workbook.active

# Cells modification
# Updating the title for the current week
worksheet[title_cell] = f'سجل الحضور والانصراف من  {start_work_day} إلى {end_work_day}'

# Updating each day column with the right date
day_index, column_increment = 0, 2
while current_columns <= 13:
    cell = worksheet.cell(row=current_rows, column=current_columns, value=dated_days[day_index])
    cell.border = Border(bottom=colored_side)
    print(f'Updated cell [{cell.coordinate}] to {dated_days[day_index]}')
    day_index+=1
    current_columns+=column_increment

# Creating a new file and saving changes into it
print(f'Saving modification into {output_file_path} file...')
workbook.save(output_file_path)
print(f'Changes saved into {output_file_path} file successfully.')